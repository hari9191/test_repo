#Importing all the required packages
import openpyxl
import tkinter as tk
from tkinter import messagebox
from tkinter import simpledialog
# from google.colab import files
# from google.colab import drive

#Assign variable with the foder path
# drive.mount('/content/drive')
# work_file1='/content/drive/MyDrive/Colab Notebooks/Worksheets/4G151_Wire_Tag_Export_Amended.xlsx'
# test_file='/content/drive/MyDrive/Colab Notebooks/Worksheets/data_test_sheet_amended.xlsx'

work_file=''
test_file='C:\\Users\\d69191\\Desktop\\Projects\\TagRename\\data_test_sheet_amended.xlsx'

#Method for search string in excel
def get_search_string():
  root=tk.Tk()
  root.withdraw()
  search_string=simpledialog.askstring("Search String", "Enter the Search String:")
  return search_string

def replace_value_in_excel(file_name, sheet_name, search_value, column_index):
  #Load the workbook
  workbook = openpyxl.load_workbook(test_file)

  #select the sheet by name
  sheet = workbook[sheet_name]

  #Iterate over the rows in the selected columns
  for row in sheet.iter_rows(min_row=2, min_col=column_index, max_col=column_index):
    for cell in row:
      #To get the corresponding cell in column 1
      corresponding_cell = sheet.cell(row=cell.row, column=1)

      #Check if the cell value matches the search value
      if cell.value is not None and str(cell.value).startswith(search_value):
        #Replace the cell value with the asterisks
        corresponding_cell.value='*'
        # cell.value = '*' * len(str(cell.value))

  #Save the modified notebook
  workbook.save(test_file)
  print("Value Replaced Successfully")

#Upload file
# uploaded=files.upload()

#Specify the file name, sheet name, column number and search value
file_name=test_file
sheet_name='Sheet1'
search_value='AF0062'
column_index=3

replace_value_in_excel(file_name, sheet_name, search_value, column_index)